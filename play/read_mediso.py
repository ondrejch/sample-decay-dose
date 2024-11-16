#!/bin/env python3
"""
Extract Bq of Ac225 from F71 file
Ondrej Chvala <ochvala@utexas.edu>
"""
import os
import re
import subprocess
import numpy as np
from datetime import datetime

NOW: str = datetime.now().replace(microsecond=0).isoformat()

SCALE_bin_path: str = os.getenv('SCALE_BIN', '/opt/scale6.3.1/bin/')
ATOM_DENS_MINIMUM: float = 0.0 # 1e-60

rad_nuclides: list = [
    'Sc-47',
    'Y-90',
    'I-123',
    'I-131',
    # 'Lu-177',
    # 'At-211',
    'Pb-212',
    'Ra-223',
    'Ac-225',
    'Th-227'
]


def get_f71_positions_index(f71file: str) -> dict:
    """ Read info of SCALE's F71 file """
    output = subprocess.run([f"{SCALE_bin_path}/obiwan", "view", "-format=info", f71file], capture_output=True)
    output = output.stdout.decode().split("\n")
    f71_idx = {}
    skip_data = ['pos', '(-)']  # sip records starting with this
    end_data = 'state definition present'  # stop reading after reaching this
    for line in output:
        data = line.split()
        if data[0].strip() in skip_data:
            continue
        if line.count(end_data) > 0:
            break
        # print(data)
        f71_idx[int(data[0])] = {'time': data[1], 'power': data[2], 'flux': data[3], 'fluence': data[4],
            'energy': data[5], 'initialhm': data[6], 'libpos': data[7], 'case': data[8], 'step': data[9],
            'DCGNAB': data[10]}
    return f71_idx


def get_last_position_for_case(f71file: str, case: int = 1) -> int:
    iii = get_f71_positions_index(f71file)
    maxi = max([i for i, rec in iii.items() if rec['case'] == str(case)])
    return maxi


def get_burned_material_atom_dens(f71file: str, position: int) -> dict:
    """ Read atom density of nuclides from SCALE's F71 file """
    output = subprocess.run(
        [f"{SCALE_bin_path}/obiwan", "view", "-format=csv", "-prec=10", "-units=atom", "-idform='{:Ee}{:AAA}{:m}'",
            f71file], capture_output=True)
    output = output.stdout.decode().split("\n")
    densities = {}  # densities[nuclide] = (density at position of f71 file)
    skip = ["case", "step", "time", "power", "flux", "volume"]
    regexp = re.compile(r"(?P<elem>[a-zA-Z]+)(?P<num>\d+)(?P<meta>m)?")
    for line in output:
        data = line.split(',')
        if data[0].strip() in skip:
            continue
        elif len(data) > 1:
            dummy = re.search(regexp, data[0].strip())
            elem = dummy.group("elem").lower()  # convert to all lower cases
            num = int(dummy.group("num"))  # to cut off leading zeros
            if dummy.group("meta"):
                nuclide = elem + "-" + str(num) + "m"  # for metastable isotopes
            else:
                nuclide = elem + "-" + str(num)
            if float(data[position]) > ATOM_DENS_MINIMUM:
                # The [x] here is what causes the code to return only the densities at position x of the f71 file
                densities[nuclide] = float(data[position])
    sorted_densities = {k: v for k, v in sorted(densities.items(), key=lambda item: -item[1])}
    return sorted_densities


def get_nuclide_Bq(f71file: str, position: int, my_nuclide: str) -> float:
    my_nuclide = my_nuclide.lower()
    output = subprocess.run(
        [f"{SCALE_bin_path}/obiwan", "view", "-format=csv", "-prec=10", "-units=becq", "-idform='{:Ee}{:AAA}{:m}'",
            f71file], capture_output=True)
    output = output.stdout.decode().split("\n")
    # bqs = {}  # densities[nuclide] = (density at position of f71 file)
    skip = ["case", "step", "time", "power", "flux", "volume"]
    regexp = re.compile(r"(?P<elem>[a-zA-Z]+)(?P<num>\d+)(?P<meta>m)?")
    for line in output:
        data = line.split(',')
        if data[0].strip() in skip:
            continue
        elif len(data) > 1:
            dummy = re.search(regexp, data[0].strip())
            elem = dummy.group("elem").lower()  # convert to all lower cases
            num = int(dummy.group("num"))  # to cut off leading zeros
            if dummy.group("meta"):
                nuclide = elem + "-" + str(num) + "m"  # for metastable isotopes
            else:
                nuclide = elem + "-" + str(num)
            if float(data[position]) > ATOM_DENS_MINIMUM:
                # The [x] here is what causes the code to return only the densities at position x of the f71 file
                # bqs[nuclide] = float(data[position])
                if nuclide == my_nuclide:
                    # print(f"{my_nuclide}: {data[position]}")
                    return float(data[position])
    # sorted_densities = {k: v for k, v in sorted(bqs.items(), key=lambda item: -item[1])}
    return -1.0


def get_nuclides_Bq(f71file: str, position: int, my_nuclides: list[str]) -> dict:
    my_nuclides = [l.lower() for l in my_nuclides]
    output = subprocess.run(
        [f"{SCALE_bin_path}/obiwan", "view", "-format=csv", "-prec=10", "-units=becq", "-idform='{:Ee}{:AAA}{:m}'",
            f71file], capture_output=True)
    output = output.stdout.decode().split("\n")
    bqs = {}  # densities[nuclide] = (density at position of f71 file)
    skip = ["case", "step", "time", "power", "flux", "volume"]
    regexp = re.compile(r"(?P<elem>[a-zA-Z]+)(?P<num>\d+)(?P<meta>m)?")
    for line in output:
        data = line.split(',')
        if data[0].strip() in skip:
            continue
        elif len(data) > 1:
            dummy = re.search(regexp, data[0].strip())
            elem = dummy.group("elem").lower()  # convert to all lower cases
            num = int(dummy.group("num"))  # to cut off leading zeros
            if dummy.group("meta"):
                nuclide = elem + "-" + str(num) + "m"  # for metastable isotopes
            else:
                nuclide = elem + "-" + str(num)
            if float(data[position]) > ATOM_DENS_MINIMUM:
                # The [x] here is what causes the code to return only the densities at position x of the f71 file
                if nuclide in my_nuclides:
                    # print(f"{my_nuclide}: {data[position]}")
                    bqs[nuclide] = float(data[position])
    sorted_bqs = {k: v for k, v in sorted(bqs.items(), key=lambda item: -item[1])}
    return sorted_bqs


def test():
    f71_file_name: str = 'ThEIRENE.f71'
    ipos: int = get_last_position_for_case(f71_file_name, 1)
    ac_225_Bq: float = get_nuclide_Bq(f71_file_name, ipos, 'Ac-225')
    print(f'Ac-225: {ac_225_Bq:.3e} Bq')  # print(f'{ac_225_Bq:.3e}')


def old():
    f71_file_name: str = 'ThEIRENE.f71'
    MTiHM: dict = {' 5.00': 10.4159200656709, '19.75': 10.2647529843048}
    # runs: dict = {
    #     '01-1_year_120_steps': {
    #         'power': 400.0, 'enr': '19.75'
    #     },
    #     '02-gasFP_removal-1_year_120_steps': {
    #         'power': 400.0, 'enr': '19.75'
    #     },
    #     '11-5pct-1_year_120_steps': {
    #         'power': 400.0, 'enr': ' 5.00'
    #     },
    #     '12-5ctp_gasFP_removal-1_year_120_steps': {
    #         'power': 400.0, 'enr': ' 5.00'
    #     },
    #     '20-20MWth/05.00pct_1y': {
    #         'power': 20.0, 'enr': ' 5.00'
    #     },
    #     '20-20MWth/05.00pct_1y_gas_removal': {
    #         'power': 20.0, 'enr': ' 5.00'
    #     },
    #     '20-20MWth/19.75pct_1y': {
    #         'power': 20.0, 'enr': '19.75'
    #     },
    #     '20-20MWth/19.75pct_1y_gas_removal': {
    #         'power': 20.0, 'enr': '19.75'
    #     },
    #     '21-01MWth/05.00pct_1y': {
    #         'power': 1.0, 'enr': ' 5.00'
    #     },
    #     '21-01MWth/05.00pct_1y_gas_removal': {
    #         'power': 1.0, 'enr': ' 5.00'
    #     },
    #     '21-01MWth/19.75pct_1y': {
    #         'power': 1.0, 'enr': '19.75'
    #     },
    #     '21-01MWth/19.75pct_1y_gas_removal':{
    #         'power': 1.0, 'enr': '19.75'
    #     },
    # }

    runs: dict = {
        '21-01MWth/05.00pct_1y': {'power':  1.0, 'enr': ' 5.00'},
        '21-01MWth/19.75pct_1y': {'power':  1.0, 'enr': '19.75'},
        '20-20MWth/05.00pct_1y': {'power': 20.0, 'enr': ' 5.00'},
        '20-20MWth/19.75pct_1y': {'power': 20.0, 'enr': '19.75'},
        '22-100MWth/05.00pct_1y': {'power': 100.0, 'enr': ' 5.00'},
        '22-100MWth/19.75pct_1y': {'power': 100.0, 'enr': '19.75'},
        '11-5pct-1_year_120_steps': {'power': 400.0, 'enr': ' 5.00'},
        '01-1_year_120_steps': {'power': 400.0, 'enr': '19.75'},
    }
    my_nuclide: str = 'Ac-225'
    print(f'{my_nuclide} production')
    for my_path, my_run in runs.items():
        my_f71_file_name = f'{my_path}/{f71_file_name}'
        thermal_power: float = my_run['power']
        enrichment: str = my_run['enr']
        ipos: int = get_last_position_for_case(my_f71_file_name, 1)
        # print(f'ipos = {ipos}')
        ac_225_Bq: float = get_nuclide_Bq(my_f71_file_name, ipos, my_nuclide)
        ac_225_Bq_pery: float = ac_225_Bq * MTiHM[enrichment]
        ac_225_Bq_pery_perMW: float = ac_225_Bq_pery / thermal_power
        doses: float = ac_225_Bq_pery_perMW / 15e3
        print(f'Power {thermal_power:5.1f} Uenr {enrichment:5s}   {ac_225_Bq_pery_perMW:.3e} Bq/y/MWth'
              f'  doses {doses:3.0f} /y/MWth')


def main():
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    import os
    import argparse
    from sample_decay_dose.isotopes import rel_iso_mass, N_A

    cwd = os.getcwd()
    parser = argparse.ArgumentParser(description='Power in MSRR mixtures.')
    parser.add_argument('--sep', required=False, default=' ', dest='sep', type=str,
                        help='record separator  for stdout')
    parser.add_argument('--xlsxname', required=False, default='mediso_ThEIRENE.xlsx',
                        dest='xlsxname', type=str, help='XLSX output file name')
    args = parser.parse_args()
    sep: str = args.sep
    xlsxname: str = args.xlsxname

    # XLS output
    wb = Workbook()
    alphabet = ' abcdefghijklmnopqrstuvwxyz'.upper()  # For number -> column
    default_font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                        strike=False, color='FF000000')
    bold_font = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False,
                     color='FF000000')
    ws = wb.create_sheet(title=xlsxname.replace('.xlsx', ''))
    ws.font = default_font
    ws.column_dimensions['A'].width *= 1.03
    ws.column_dimensions['B'].width *= 1.05

    # Title
    ws['A1'] = f'Medical isotope production from Th-EIRENE model'
    ws['A1'].font = bold_font
    ws['A2'] = NOW
    # Header Bq
    irow: int = 3  # starting row #
    ws[f'C{irow}'] = 'Activity [Bq / MWth] in year 1'
    ws[f'C{irow}'].alignment = Alignment(horizontal='center')
    last_col = alphabet[len(rad_nuclides) + 2]
    ws.merge_cells(f'C{irow}:{last_col}{irow}')
    # Header U233
    col1: str = alphabet[len(rad_nuclides) + 4]
    col2: str = alphabet[len(rad_nuclides) + 5]
    col3: str = alphabet[len(rad_nuclides) + 6]
    ws[f'{col1}{irow}'] = 'Total U-233 after 1 year'
    ws[f'{col1}{irow}'].alignment = Alignment(horizontal='center')
    ws.merge_cells(f'{col1}{irow}:{col3}{irow}')
    ws[f'{col1}{irow + 1}'] = '[atoms]'
    ws[f'{col2}{irow + 1}'] = '[grams]'
    ws[f'{col3}{irow + 1}'] = '[g / MW]'
    # Header Bq
    irow += 1
    ws[f'A{irow}'] = 'Fuel type'
    ws[f'B{irow}'] = 'Power [MWth]'
    for i, nuc in enumerate(rad_nuclides):
        col = alphabet[i+3]
        ws[f'{col}{irow}'] = nuc
    for cc in ws[f'C{irow}:{col3}{irow}']:
        for c in cc:
            c.alignment = Alignment(horizontal='center')

    # Read F71 files
    f71_file_name: str = 'ThEIRENE.f71'
    fuel_volume: float = 1.06806e+07  # [cm^3]
    MTiHM: dict = {' 5.00': 10.4159200656709, '19.75': 10.2647529843048}
    fuel_type: dict = {' 5.00': 'LEU+Th', '19.75': 'HALEU+Th'}
    runs: dict = {
        '21-01MWth/05.00pct_1y': {'power':  1.0, 'enr': ' 5.00'},
        '21-01MWth/19.75pct_1y': {'power':  1.0, 'enr': '19.75'},
        '20-20MWth/05.00pct_1y': {'power': 20.0, 'enr': ' 5.00'},
        '20-20MWth/19.75pct_1y': {'power': 20.0, 'enr': '19.75'},
        '22-100MWth/05.00pct_1y': {'power': 100.0, 'enr': ' 5.00'},
        '22-100MWth/19.75pct_1y': {'power': 100.0, 'enr': '19.75'},
        # '11-5pct-1_year_120_steps': {'power': 400.0, 'enr': ' 5.00'},
        # '01-1_year_120_steps': {'power': 400.0, 'enr': '19.75'},
    }
    print(f'{rad_nuclides} production')

    for my_path, my_run in runs.items():
        my_f71_file_name = f'{my_path}/{f71_file_name}'
        thermal_power: float = my_run['power']
        enrichment: str = my_run['enr']
        ipos: int = get_last_position_for_case(my_f71_file_name, 1)
        # print(f'ipos = {ipos}')
        activities_pery: dict = get_nuclides_Bq(my_f71_file_name, ipos, rad_nuclides)
        activities_pery = {nuc: activities_pery[nuc] * MTiHM[enrichment] for nuc in activities_pery.keys()}
        print(activities_pery)
        activities_pery_perMW: dict = {nuc: activities_pery[nuc] / thermal_power for nuc in activities_pery.keys()}
        print(activities_pery_perMW)
        # print(f'Power {thermal_power:5.1f} Uenr {enrichment:5s}   {ac_225_Bq_pery_perMW:.3e} Bq/y/MWth  doses {doses:3.0f} /y/MWth')
        irow += 1
        ws[f'A{irow}'] = fuel_type[enrichment]
        ws[f'B{irow}'].alignment = Alignment(horizontal='center')
        ws[f'B{irow}'] = thermal_power
        ws[f'B{irow}'].alignment = Alignment(horizontal='center')
        for i, nuc in enumerate(rad_nuclides):
            col = alphabet[i + 3]
            ws[f'{col}{irow}'] = activities_pery_perMW[nuc.lower()]
            ws[f'{col}{irow}'].number_format = '0.00e+00'

        at_dens: dict = get_burned_material_atom_dens(my_f71_file_name, ipos)
        u233_atoms: float = at_dens['u-233'] * fuel_volume * 1e24
        u233_grams: float = u233_atoms * rel_iso_mass['u-233'] / N_A
        col = alphabet[len(rad_nuclides) + 4]
        ws[f'{col}{irow}'] = u233_atoms
        ws[f'{col}{irow}'].number_format = '0.00e+00'
        col = alphabet[len(rad_nuclides) + 5]
        ws[f'{col}{irow}'] = u233_grams
        ws[f'{col}{irow}'].number_format = '0.000e+0'
        col = alphabet[len(rad_nuclides) + 6]
        ws[f'{col}{irow}'] = u233_grams / thermal_power
        ws[f'{col}{irow}'].number_format = '0.0'

    irow += 2
    ws[f'A{irow}'] = 'LEU+Th'
    ws[f'B{irow}'] = '61.333%LiF + 30.667%BeF2 + 5.300%UF4 + 2.700%ThF4, U enrichment 0.05'
    irow += 1
    ws[f'A{irow}'] = 'HALEU+Th'
    ws[f'B{irow}'] = '61.333%LiF + 30.667%BeF2 + 1.420%UF4 + 6.580%ThF4, U enrichment 0.1975'
    irow += 2
    ws[f'A{irow}'] = 'Sc-47'
    ws[f'B{irow}'] = 'Produced by target irradiation, see https://www.sciencedirect.com/science/article/abs/pii/S0969804316306649'
    irow += 1
    ws[f'A{irow}'] = 'Lu-177'
    ws[f'B{irow}'] = 'Produced by target irradiation, see https://www.ncbi.nlm.nih.gov/pmc/articles/PMC4463871/'
    irow += 1
    ws[f'A{irow}'] = 'At-211'
    ws[f'B{irow}'] = 'Produced by target irradiation, see https://www.ncbi.nlm.nih.gov/pmc/articles/PMC3503149/'

    if 'Sheet' in wb.sheetnames:  # Delete default page if exists
        wb.remove(wb['Sheet'])
    wb.save(xlsxname)  # Save XLS file


if __name__ == "__main__":
    # old()
    main()
